import tkinter as tk
from tkinter import ttk, messagebox
import requests
from bs4 import BeautifulSoup
import pyttsx3
from deep_translator import GoogleTranslator
import google.generativeai as genai
import threading
import json
import os
from datetime import datetime
import time
import sys
import bisect
import hashlib
from pathlib import Path

from dotenv import load_dotenv, dotenv_values
from io import StringIO


def load_environment():
    """Load environment variables from common locations.

    Khi đóng gói bằng PyInstaller, __file__ có thể khác hoặc không tồn tại và
    .env có thể nằm cùng thư mục với file .exe. Hàm này sẽ dò nhiều vị trí
    khả dụng để đảm bảo biến môi trường được nạp."""

    search_paths = []

    # PyInstaller unpacked temp directory (sys._MEIPASS)
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        search_paths.append(Path(meipass))

    # Thư mục chứa file thực thi khi chạy dưới dạng .exe
    if getattr(sys, "frozen", False):
        search_paths.append(Path(sys.executable).resolve().parent)

    # Thư mục chứa file source và thư mục hiện tại
    search_paths.append(Path(__file__).resolve().parent)
    search_paths.append(Path.cwd())

    loaded = False
    for path in search_paths:
        env_path = path / ".env"
        try:
            if env_path.exists():
                # Lần đầu tiên load sẽ override các biến hiện tại
                load_result = load_dotenv(env_path, override=not loaded)
                if load_result:
                    loaded = True
                    print(f"[Env] Loaded .env from {env_path}")
                else:
                    # Khi file .env sử dụng encoding khác UTF-8 (ví dụ UTF-16) thì
                    # load_dotenv trả về False mà không raise exception. Thử parse
                    # thủ công với một vài encoding phổ biến để tránh lỗi khi đóng gói.
                    raw_bytes = env_path.read_bytes()
                    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be"):
                        try:
                            content = raw_bytes.decode(encoding)
                        except UnicodeDecodeError:
                            continue
                        values = dotenv_values(stream=StringIO(content))
                        if values:
                            for key, value in values.items():
                                if value is None:
                                    continue
                                if not loaded or key not in os.environ:
                                    os.environ[key] = value
                            loaded = True
                            print(
                                f"[Env] Loaded .env from {env_path} using fallback encoding {encoding}"
                            )
                            break
        except UnicodeDecodeError as e:
            print(f"[Env] Could not decode {env_path}: {e}")
        except Exception as e:
            print(f"[Env] Could not load {env_path}: {e}")

    # Fallback để đảm bảo biến môi trường hệ thống vẫn được đọc
    if not loaded:
        load_dotenv()


# load_environment()  # Disabled to avoid .env encoding issues

class CambridgeDictionaryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cambridge Dictionary - AI Enhanced")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        # Configure modern styling
        self.root.configure(bg="#f8f9fa")
        
        # Khởi tạo TTS engine - Chỉ tắt âm thanh tác vụ
        try:
            self.tts_engine = pyttsx3.init()
            # Tắt âm thanh click button và tác vụ
            self.tts_engine.setProperty('volume', 0.0)
        except:
            self.tts_engine = None
        
        # Translator với retry logic
        self.translator = GoogleTranslator(source='en', target='vi')
        
        # API Key Rotation System - Multiple API keys for extended quota
        self.api_keys = [
            'AIzaSyCz0JtTfcbSjhQ54wux1QPHvQGDGCjbzmw',  # Key 1
            'AIzaSyCaAHdPoLt4ZiJMv3Kv5b6-X9UWrkNCYDo',  # Key 2
            'AIzaSyCz7CazBX2o2-hfBag4pQuKYJXOqKTgrm8',  # Key 3
            'AIzaSyAe7C3CSmAb8j7NuqZSxXJpWO3DITUaqZA',  # Key 4
            'AIzaSyA-fqbIETEndKrMmaT83GEFvFNQ1STmOxQ',  # Key 5
            # Có thể thêm nhiều keys hơn tại đây
        ]
        self.current_key_index = 0
        self.gemini_enabled = False
        self.gemini_model = None
        
        # Initialize with first available key
        self._initialize_gemini()

        # Cache để tăng tốc EXTREME
        self.translation_cache = {}
        self.word_cache = {}  # Cache cả từ đã tra

        # Audio cache
        self.audio_cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'audio_cache')
        try:
            os.makedirs(self.audio_cache_dir, exist_ok=True)
        except Exception as e:
            print(f"[Audio Cache] Could not create cache directory: {e}")
            self.audio_cache_dir = None
        
        # File lưu từ vựng
        self.vocab_file = "my_vocabulary.json"
        self.load_vocabulary()
        
        # Session để tái sử dụng kết nối
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        # Load cache
        self.load_cache()
        
        # Thiết lập giao diện
        self.setup_ui()
        
    def load_cache(self):
        """Load translation cache từ file"""
        try:
            if os.path.exists('translation_cache.json'):
                with open('translation_cache.json', 'r', encoding='utf-8') as f:
                    self.translation_cache = json.load(f)
        except:
            pass
    
    def save_cache(self):
        """Lưu translation cache"""
        try:
            with open('translation_cache.json', 'w', encoding='utf-8') as f:
                json.dump(self.translation_cache, f, ensure_ascii=False, indent=2)
        except:
            pass
        
    def load_vocabulary(self):
        """Load từ vựng đã lưu"""
        try:
            if os.path.exists(self.vocab_file):
                with open(self.vocab_file, 'r', encoding='utf-8') as f:
                    self.vocabulary = json.load(f)
            else:
                self.vocabulary = []
        except:
            self.vocabulary = []
    
    def save_vocabulary(self):
        """Lưu từ vựng"""
        try:
            with open(self.vocab_file, 'w', encoding='utf-8') as f:
                json.dump(self.vocabulary, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Lỗi khi lưu từ vựng: {e}")
    
    def add_to_vocabulary(self, word, phonetic_uk, phonetic_us, definitions, ai_vi=None, ai_example_en=None):
        """Thêm từ vào danh sách ghi nhớ"""
        vocab_item = {
            'word': word,
            'phonetic_uk': phonetic_uk,
            'phonetic_us': phonetic_us,
            'definitions': definitions,
            'added_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        if ai_vi:
            vocab_item['ai_vi'] = ai_vi
        if ai_example_en:
            vocab_item['ai_example_en'] = ai_example_en
        
        for item in self.vocabulary:
            if item['word'].lower() == word.lower():
                # Update existing instead of duplicate
                item['phonetic_uk'] = phonetic_uk
                item['phonetic_us'] = phonetic_us
                item['definitions'] = definitions
                if ai_vi:
                    item['ai_vi'] = ai_vi
                if ai_example_en:
                    item['ai_example_en'] = ai_example_en
                self.save_vocabulary()
                self.show_toast(f"Đã cập nhật '{word}' trong danh sách!")
                return
        
        self.vocabulary.append(vocab_item)
        self.save_vocabulary()
        self.show_toast(f"Đã thêm '{word}' vào danh sách!")
    
    def show_toast(self, message):
        """Hiển thị thông báo toast 1s tự đóng - cố định ở giữa màn hình"""
        toast = tk.Toplevel(self.root)
        toast.title("")
        toast.geometry("350x80")
        toast.resizable(False, False)
        
        # Cố định ở giữa màn hình
        screen_width = toast.winfo_screenwidth()
        screen_height = toast.winfo_screenheight()
        x = (screen_width - 350) // 2
        y = (screen_height - 80) // 2
        toast.geometry(f"350x80+{x}+{y}")
        
        # Làm cho cửa sổ luôn ở trên
        toast.attributes('-topmost', True)
        toast.overrideredirect(True)  # Bỏ title bar
        
        # Tạo frame với màu nền
        toast_frame = tk.Frame(toast, bg=self.colors['accent'], relief=tk.RAISED, bd=2)
        toast_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Label thông báo
        toast_label = tk.Label(
            toast_frame,
            text=message,
            font=("Segoe UI", 13, "bold"),
            bg=self.colors['accent'],
            fg=self.colors['white'],
            wraplength=320,
            justify=tk.CENTER
        )
        toast_label.pack(expand=True)
        
        # Tự đóng sau 1s
        toast.after(1000, toast.destroy)
        
    def setup_ui(self):
        # Modern color scheme
        self.colors = {
            'primary': '#1e3a8a',      # Deep blue
            'secondary': '#3b82f6',    # Bright blue
            'accent': '#10b981',       # Green
            'warning': '#f59e0b',      # Orange
            'danger': '#ef4444',       # Red
            'dark': '#1f2937',         # Dark gray
            'light': '#f8fafc',        # Light gray
            'light_gray': '#e5e7eb',   # Light gray for separators
            'white': '#ffffff',
            'border': '#e5e7eb'        # Light border
        }
        
        # Header with modern design
        self.setup_header()
        
        # Tab system
        self.setup_tab_system()
        
        # Search bar
        self.setup_search_bar()
        
        # Main content area
        self.setup_main_content()

        # Create initial status
        self.update_current_word_display("")
        self.show_welcome()

    def update_current_word_display(self, word):
        display = word.strip() if word else ""
        if hasattr(self, 'current_word_var'):
            self.current_word_var.set(f"{display}")
        
    def setup_header(self):
        """Setup modern header"""
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Title with icon
        title_frame = tk.Frame(header_frame, bg=self.colors['primary'])
        title_frame.pack(side=tk.LEFT, padx=20, pady=15)
        
        title_label = tk.Label(
            title_frame,
            text="Cambridge Dictionary",
            font=("Segoe UI", 20, "bold"),
            bg=self.colors['primary'],
            fg=self.colors['white']
        )
        title_label.pack(side=tk.LEFT, padx=(0, 20))
        
        # Dictionary search bar (like Cambridge Dictionary) - wider
        search_frame = tk.Frame(header_frame, bg=self.colors['white'], relief=tk.SOLID, borderwidth=1)
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(20, 20))
        
        self.search_entry = tk.Entry(
            search_frame,
            font=("Segoe UI", 12),
            bg=self.colors['white'],
            relief=tk.FLAT,
            borderwidth=0
        )
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=15, pady=8)
        self.search_entry.bind('<Return>', self.on_search_enter)
        
        # Bind suggestions to search_entry
        self.search_entry.bind('<KeyRelease>', self.on_key_release)
        self.search_entry.bind('<Down>', self.focus_suggestion)
        self.search_entry.bind('<Up>', self.focus_suggestion_up)
        self.search_entry.bind('<Escape>', lambda e: self.hide_suggestions())
        
        # Bind paste để tự động gợi ý sau khi paste
        self.search_entry.bind('<Control-v>', self.on_paste)
        self.search_entry.bind('<Button-3>', self.on_right_click)
        
        # Search button
        search_btn = tk.Button(
            search_frame,
            text="🔍",
            font=("Segoe UI", 16),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor="hand2",
            command=self.search_word
        )
        search_btn.pack(side=tk.RIGHT)
        
        # Right side buttons
        btn_frame = tk.Frame(header_frame, bg=self.colors['primary'])
        btn_frame.pack(side=tk.RIGHT, padx=20, pady=15)
        
        # Text Translator button
        translator_btn = tk.Button(
            btn_frame,
            text="Dịch văn bản",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=15,
            pady=5,
            cursor="hand2",
            command=self.show_text_translator
        )
        translator_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Vocabulary button
        vocab_btn = tk.Button(
            btn_frame,
            text="Vocabulary",
            font=("Segoe UI", 10, "bold"),
            bg=self.colors['accent'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=15,
            pady=5,
            cursor="hand2",
            command=self.show_vocabulary
        )
        vocab_btn.pack(side=tk.LEFT)
        
    def setup_tab_system(self):
        """Setup single-page status bar"""
        self.tab_frame = tk.Frame(self.root, bg=self.colors['light'], height=30)
        self.tab_frame.pack(fill=tk.X, padx=20, pady=(5, 0))
        self.tab_frame.pack_propagate(False)

        self.current_word_var = tk.StringVar(value="")
        self.current_word_label = tk.Label(
            self.tab_frame,
            textvariable=self.current_word_var,
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['light'],
            fg=self.colors['dark']
        )
        self.current_word_label.pack(side=tk.LEFT, padx=10)
        
    def setup_search_bar(self):
        """Setup suggestions listbox - positioned in header below search bar"""
        # Suggestion listbox (hidden initially) - will be positioned in header
        self.suggestion_listbox = tk.Listbox(
            self.root,
            font=("Segoe UI", 11),
            height=5,
            relief=tk.SOLID,
            bd=1,
            bg=self.colors['white'],
            selectbackground=self.colors['secondary']
        )
        # Don't pack yet - will be positioned dynamically
        self.suggestion_listbox.bind('<<ListboxSelect>>', self.on_suggestion_select)
        self.suggestion_listbox.bind('<Button-1>', self.on_suggestion_click)
        self.suggestion_listbox.bind('<Return>', self.on_suggestion_enter)
        self.suggestion_listbox.bind('<Up>', self.on_suggestion_up)
        self.suggestion_listbox.bind('<Down>', self.on_suggestion_down)
        self.suggestion_listbox.bind('<Escape>', lambda e: self.hide_suggestions())
        
        # History và common words
        self.search_history = []
        self.common_words = self.load_common_words()
        
    def setup_main_content(self):
        """Setup main content area"""
        # Main content frame
        self.main_frame = tk.Frame(self.root, bg=self.colors['white'])
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
        
        # Content area for current tab
        self.content_frame = tk.Frame(self.main_frame, bg=self.colors['white'])
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Canvas for scrolling
        self.result_canvas = tk.Canvas(self.content_frame, bg=self.colors['white'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.content_frame, orient="vertical", command=self.result_canvas.yview)
        self.scrollable_frame = tk.Frame(self.result_canvas, bg=self.colors['white'])
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.result_canvas.configure(scrollregion=self.result_canvas.bbox("all"))
        )
        
        self.result_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.result_canvas.configure(yscrollcommand=scrollbar.set)
        
        self.result_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.result_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # Context + AI controls area
        context_section = tk.Frame(self.main_frame, bg=self.colors['white'])
        context_section.pack(fill=tk.X, pady=(10, 0))

        # Context checkbox and label
        context_header_frame = tk.Frame(context_section, bg=self.colors['white'])
        context_header_frame.pack(fill=tk.X)
        
        self.use_context_var = tk.BooleanVar()
        context_checkbox = tk.Checkbutton(
            context_header_frame,
            text="Sử dụng ngữ cảnh để dịch chính xác hơn",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark'],
            variable=self.use_context_var,
            command=self._toggle_context_input
        )
        context_checkbox.pack(side=tk.LEFT)
        
        tk.Label(
            context_header_frame,
            text="Ngữ cảnh (tùy chọn) cho bản dịch AI:",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(side=tk.LEFT, padx=(20, 0))

        self.context_text = tk.Text(
            context_section,
            height=3,
            font=("Segoe UI", 11),
            wrap=tk.WORD,
            bg=self.colors['light'],
            relief=tk.SOLID,
            borderwidth=1,
            padx=8,
            pady=6
        )
        # Initially hidden - only show when checkbox is ticked

        ai_control_frame = tk.Frame(context_section, bg=self.colors['white'])
        ai_control_frame.pack(fill=tk.X)

        self.ai_status_var = tk.StringVar()
        self.ai_status_label = tk.Label(
            ai_control_frame,
            textvariable=self.ai_status_var,
            font=("Segoe UI", 10),
            bg=self.colors['white'],
            fg=self.colors['secondary']
        )
        self.ai_status_label.pack(side=tk.LEFT, anchor=tk.W)

        self.ai_translate_btn = tk.Button(
            ai_control_frame,
            text="AI dịch theo ngữ cảnh",
            font=("Segoe UI", 11, "bold"),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=14,
            pady=6,
            cursor="hand2",
            command=self.run_ai_translate_current
        )
        self.ai_translate_btn.pack(side=tk.RIGHT)

        if not self.gemini_enabled:
            self.ai_translate_btn.configure(state=tk.DISABLED, bg=self.colors['border'], fg=self.colors['dark'])

        self.reset_ai_ui_state(reset_context=True)

    def reset_ai_ui_state(self, reset_context=False, clear_word_info=True):
        """Reset trạng thái liên quan đến AI/context"""
        self.current_ai_vi = None
        self.current_ai_example_en = None
        self.ai_vi_label = None
        if reset_context and hasattr(self, 'context_text') and self.context_text.winfo_exists():
            self.context_text.delete('1.0', tk.END)
        if hasattr(self, 'ai_status_var'):
            default_status = (
                "AI chưa khả dụng - kiểm tra GEMINI_API_KEY"
                if not self.gemini_enabled
                else "Nhập ngữ cảnh tùy chọn rồi bấm nút để dịch"
            )
            self.ai_status_var.set(default_status)
        if hasattr(self, 'ai_translate_btn'):
            btn_state = tk.NORMAL if self.gemini_enabled else tk.DISABLED
            btn_bg = self.colors['secondary'] if self.gemini_enabled else self.colors['border']
            btn_fg = self.colors['white'] if self.gemini_enabled else self.colors['dark']
            self.ai_translate_btn.configure(state=btn_state, text="🤖 AI dịch theo ngữ cảnh", bg=btn_bg, fg=btn_fg)
        if clear_word_info:
            self.current_word_info = None

    def _on_mousewheel(self, event):
        self.result_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
    def show_welcome(self):
        """Show empty state - no welcome message needed"""
        self.update_current_word_display("")
        self.reset_ai_ui_state(reset_context=True)
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        # Just show empty state - no welcome message
        
    def clear_results(self):
        self.reset_ai_ui_state(clear_word_info=True)
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

    def translate_text(self, text):
        """Dịch với AI (ưu tiên) hoặc Google Translate fallback"""
        # Kiểm tra cache trước
        if text in self.translation_cache:
            return self.translation_cache[text]

        # Ưu tiên dùng AI nếu enabled
        vi_text = None
        if self.gemini_enabled:
            try:
                vi_text = self.ai_translate_simple(text)
            except Exception as e:
                print(f"[AI Translate] Error: {e}")

        if not vi_text:
            vi_text = self._google_translate_text(text)

        if vi_text:
            self.translation_cache[text] = vi_text
            if len(self.translation_cache) % 10 == 0:
                self.save_cache()
            return vi_text

        return text

    def _run_batch_translation(self, translation_targets):
        def worker():
            self._process_batch_translations(translation_targets)

        threading.Thread(target=worker, daemon=True).start()

    def _process_batch_translations(self, translation_targets):
        pending_items = [t for t in translation_targets if t.get('text')]
        if not pending_items:
            return

        resolved = {}
        uncached_texts = []

        for item in pending_items:
            text = item['text']
            if text in self.translation_cache:
                resolved[text] = self.translation_cache[text]
            else:
                uncached_texts.append(text)

        unique_uncached = list(dict.fromkeys(uncached_texts))
        ai_results = {}

        if unique_uncached and self.gemini_enabled:
            try:
                ai_results = self.ai_translate_batch(unique_uncached)
            except Exception as e:
                print(f"[AI Batch] Error: {e}")

        for text, translation in ai_results.items():
            if translation:
                resolved[text] = translation
                self.translation_cache[text] = translation

        remaining = [t for t in unique_uncached if t not in resolved]
        for text in remaining:
            translation = self._google_translate_text(text)
            if translation:
                resolved[text] = translation
                self.translation_cache[text] = translation

        if self.translation_cache and len(self.translation_cache) % 10 == 0:
            self.save_cache()

        for item in pending_items:
            text = item['text']
            label = item['label']
            prefix = item.get('prefix', '')
            translation = resolved.get(text)
            if not translation:
                translation = text

            display_text = f"{prefix}{translation}"
            self.root.after(0, lambda lbl=label, txt=display_text: self._safe_update_label(lbl, txt))
    
    def _safe_update_label(self, label, text):
        """Cập nhật label một cách an toàn"""
        try:
            if label and label.winfo_exists():
                label.config(text=text)
        except:
            pass

    def _google_translate_text(self, text):
        for attempt in range(2):
            try:
                translation = self.translator.translate(text)
                return translation
            except Exception as e:
                if attempt == 0:
                    time.sleep(0.5)
                    continue
                print(f"[Google Translate] Error: {e}")
        return None
    
    def ai_translate_simple(self, text: str) -> str:
        """Dịch đơn giản EN->VI bằng AI"""
        try:
            prompt = (
                f"Translate the following English text to Vietnamese. Use the simplest, "
                f"most common, and easiest-to-understand meaning. Return only the translation: {text}"
            )
            resp = self.gemini_model.generate_content(prompt)
            result = (getattr(resp, 'text', None) or '').strip()
            # Remove quotes if wrapped
            if result.startswith('"') and result.endswith('"'):
                result = result[1:-1]
            if result.startswith("'") and result.endswith("'"):
                result = result[1:-1]
            return result if result else None
        except Exception as e:
            print(f"[AI Simple] Error: {e}")
            return None

    def ai_translate_batch(self, texts):
        """Dịch nhiều đoạn văn bản trong một lần gọi AI"""
        if not texts:
            return {}

        # Loại bỏ trùng lặp nhưng giữ thứ tự
        unique_texts = list(dict.fromkeys(texts))
        payload = json.dumps(unique_texts, ensure_ascii=False)
        prompt = (
            "Translate each English string in the JSON array below to Vietnamese. "
            "Respond ONLY with a JSON array of translations, keeping the order identical to the input.\n\n"
            f"Input:\n{payload}\n\nOutput:\n"
        )

        try:
            resp = self.gemini_model.generate_content(prompt)
            raw_text = (getattr(resp, 'text', None) or '').strip()

            if raw_text.startswith('```'):
                parts = raw_text.split('```')
                for part in parts:
                    if part.strip().startswith('{') or part.strip().startswith('['):
                        raw_text = part.strip()
                        break

            import re as _re

            match = _re.search(r"\[[\s\S]*\]", raw_text)
            if match:
                raw_text = match.group(0)

            translations = json.loads(raw_text)
            if not isinstance(translations, list):
                return {}

            result = {}
            for original, translated in zip(unique_texts, translations):
                if isinstance(translated, str) and translated.strip():
                    result[original] = translated.strip().strip('"').strip("'")
            return result
        except Exception as e:
            print(f"[AI Batch] Parse error: {e}")
            return {}
    
    def search_word(self, event=None):
        word = self.search_entry.get().strip()

        if not word:
            self.show_toast("Please enter a word!")
            return

        # Update status bar for this word
        self.update_current_word_display(word)

        # Ẩn danh sách gợi ý khi bắt đầu tra cứu
        self.hide_suggestions()

        # Clear search entry
        self.search_entry.delete(0, tk.END)

        # Kiểm tra word cache
        if word.lower() in self.word_cache:
            print(f"⚡ Loading from cache: {word}")
            self.root.after(0, lambda: self._display_results(self.word_cache[word.lower()]))
            return
        
        self.clear_results()
        loading_label = tk.Label(
            self.scrollable_frame,
            text=f"Searching for '{word}'...",
            font=("Segoe UI", 16),
            bg=self.colors['white'],
            fg=self.colors['dark']
        )
        loading_label.pack(pady=50)
        
        # Add loading animation
        loading_dots = tk.Label(
            self.scrollable_frame,
            text="⏳",
            font=("Segoe UI", 20),
            bg=self.colors['white'],
            fg=self.colors['secondary']
        )
        loading_dots.pack()
        
        threading.Thread(target=self._search_word_thread, args=(word,), daemon=True).start()

    def _search_word_thread(self, word):
        start_time = time.time()
        word_info = self.get_word_info(word)
        elapsed = time.time() - start_time
        print(f"⏱️ Search time: {elapsed:.2f}s")
        
        if not word_info:
            self.root.after(0, lambda: self._show_error(word))
            return
        
        # Lưu vào word cache
        self.word_cache[word.lower()] = word_info
        
        # Thêm vào history (không trùng lặp)
        if word.lower() not in [w.lower() for w in self.search_history]:
            self.search_history.insert(0, word.lower())
            if len(self.search_history) > 50:  # Giữ tối đa 50 từ
                self.search_history.pop()
        
        # Tự động thêm từ mới vào common_words.txt nếu chưa có
        self.add_word_to_suggestions(word)
        
        self.root.after(0, lambda: self._display_results(word_info))
    
    def on_key_release(self, event):
        """Xử lý khi user gõ - hiển thị suggestions"""
        # Bỏ qua các phím đặc biệt
        if event.keysym in ['Return', 'Up', 'Down', 'Left', 'Right', 'Escape']:
            if event.keysym == 'Escape':
                self.hide_suggestions()
            return
        
        text = self.search_entry.get().strip().lower()
        
        if len(text) < 2:  # Chỉ suggest khi gõ từ 2 ký tự trở lên
            self.hide_suggestions()
            return
        
        # Tìm từ gợi ý
        suggestions = self.get_suggestions(text)
        
        if suggestions:
            self.show_suggestions(suggestions)
        else:
            self.hide_suggestions()
    
    def on_paste(self, event):
        """Xử lý khi paste (Ctrl+V) - tự động gợi ý sau khi paste"""
        # Chờ paste hoàn thành
        self.root.after(50, self._refresh_suggestions_after_paste)
        return None  # Allow default paste behavior
    
    def on_right_click(self, event):
        """Xử lý khi click chuột phải"""
        # Chờ paste hoàn thành (nếu user chọn Paste từ menu)
        self.root.after(100, self._refresh_suggestions_after_paste)
        return None
    
    def _refresh_suggestions_after_paste(self):
        """Refresh suggestions sau khi paste"""
        text = self.search_entry.get().strip().lower()
        if text and len(text) >= 2:
            suggestions = self.get_suggestions(text)
            if suggestions:
                self.show_suggestions(suggestions)
            else:
                self.hide_suggestions()
    
    def add_word_to_suggestions(self, word):
        """Thêm từ mới vào common_words.txt"""
        if not word or len(word) < 2:
            return
        
        word_lower = word.lower().strip()
        
        try:
            # Xác định đường dẫn file
            if getattr(sys, 'frozen', False):
                # Chạy từ .exe - lưu vào thư mục exe
                base_path = Path(sys.executable).resolve().parent
            else:
                # Chạy từ .py
                base_path = Path(__file__).resolve().parent
            
            file_path = base_path / 'common_words.txt'
            
            # Kiểm tra xem từ đã tồn tại chưa
            existing_words = []
            if file_path.exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_words = [line.strip().lower() for line in f if line.strip()]
            
            # Nếu từ chưa tồn tại, thêm vào
            if word_lower not in existing_words:
                with open(file_path, 'a', encoding='utf-8') as f:
                    f.write(f"{word_lower}\n")
                print(f"✅ Added '{word_lower}' to common_words.txt")
                # Reload common words
                self.common_words = self.load_common_words()
        except Exception as e:
            print(f"❌ Error adding word to suggestions: {e}")
    
    def load_common_words(self):
        """Load danh sách từ phổ biến từ file common_words.txt"""
        try:
            # Xác định đường dẫn file
            if getattr(sys, 'frozen', False):
                # Chạy từ .exe (PyInstaller)
                base_path = sys._MEIPASS
            else:
                # Chạy từ .py
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            file_path = os.path.join(base_path, 'common_words.txt')
            
            # Đọc file
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    words = [line.strip().lower() for line in f if line.strip()]
                print(f"✅ Loaded {len(words)} common words from file")
                return sorted(set(words))  # Remove duplicates và sort
            else:
                print(f"⚠️ File not found: {file_path}")
                return self.get_fallback_words()
        except Exception as e:
            print(f"❌ Error loading common_words.txt: {e}")
            return self.get_fallback_words()
    
    def get_fallback_words(self):
        """Fallback words nếu không load được file"""
        return sorted([
            'hello', 'world', 'computer', 'beautiful', 'wonderful',
            'amazing', 'excellent', 'important', 'different', 'interesting',
            'development', 'education', 'environment', 'government',
            'information', 'technology', 'university', 'communication'
        ])
    
    def get_suggestions(self, text):
        """Lấy danh sách từ gợi ý - OPTIMIZED với Binary Search"""
        suggestions = []
        
        # 1. Từ trong vocabulary (ưu tiên cao nhất)
        for item in self.vocabulary:
            word = item['word'].lower()
            if word.startswith(text):
                suggestions.append(word)
        
        # 2. Từ trong history
        for word in self.search_history:
            if word.startswith(text) and word not in suggestions:
                suggestions.append(word)
        
        # 3. Từ phổ biến (Binary Search cho speed!)
        if self.common_words:
            # Tìm vị trí bắt đầu với binary search
            idx = bisect.bisect_left(self.common_words, text)
            
            # Lấy tối đa 50 từ để check
            for i in range(idx, min(idx + 50, len(self.common_words))):
                word = self.common_words[i]
                if word.startswith(text):
                    if word not in suggestions:
                        suggestions.append(word)
                else:
                    break  # Không match nữa thì dừng
        
        # Giới hạn 10 suggestions, sort theo alphabet
        return sorted(set(suggestions))[:10]
    
    def show_suggestions(self, suggestions):
        """Hiển thị dropdown suggestions ngay dưới search bar, căn giữa như Cambridge Dictionary"""
        self.suggestion_listbox.delete(0, tk.END)
        
        for suggestion in suggestions:
            self.suggestion_listbox.insert(tk.END, suggestion)
        
        # Position suggestions ngay dưới search bar, căn giữa với search bar
        header_frame = self.root.winfo_children()[0]  # Header là child đầu tiên
        self.suggestion_listbox.pack(fill=tk.X, padx=20, pady=(0, 5), after=header_frame)
        
        # Select first item for keyboard nav
        if self.suggestion_listbox.size() > 0:
            self.suggestion_listbox.selection_set(0)
    
    def hide_suggestions(self):
        """Ẩn dropdown suggestions"""
        self.suggestion_listbox.pack_forget()
    
    def on_suggestion_select(self, event):
        """Khi chọn suggestion bằng keyboard"""
        selection = self.suggestion_listbox.curselection()
        if selection:
            word = self.suggestion_listbox.get(selection[0])
            self.search_entry.delete(0, tk.END)
            self.search_entry.insert(0, word)
    
    def on_suggestion_click(self, event):
        """Khi click vào suggestion"""
        try:
            if self.suggestion_listbox.size() == 0:
                return

            index = self.suggestion_listbox.nearest(event.y)
            if index >= 0:
                self.suggestion_listbox.selection_clear(0, tk.END)
                self.suggestion_listbox.selection_set(index)
                word = self.suggestion_listbox.get(index)
                self.search_entry.delete(0, tk.END)
                self.search_entry.insert(0, word)
                self.hide_suggestions()
                self.search_word()
        except:
            pass

    def focus_suggestion(self, event):
        try:
            if self.suggestion_listbox.size() > 0:
                self.suggestion_listbox.focus_set()
                self.suggestion_listbox.activate(0)
                self.suggestion_listbox.selection_clear(0, tk.END)
                self.suggestion_listbox.selection_set(0)
        except:
            pass
        return 'break'

    def on_suggestion_enter(self, event):
        try:
            selection = self.suggestion_listbox.curselection()
            if selection:
                word = self.suggestion_listbox.get(selection[0])
                self.search_entry.delete(0, tk.END)
                self.search_entry.insert(0, word)
                self.hide_suggestions()
                self.search_word()
        except:
            pass
        return 'break'
    
    def on_search_enter(self, event):
        """Xử lý Enter trong search entry - chọn gợi ý đầu tiên nếu có"""
        if self.suggestion_listbox.winfo_viewable() and self.suggestion_listbox.size() > 0:
            # Chọn gợi ý đầu tiên
            word = self.suggestion_listbox.get(0)
            self.search_entry.delete(0, tk.END)
            self.search_entry.insert(0, word)
            self.hide_suggestions()
            self.search_word()
        else:
            # Tra từ hiện tại
            self.search_word()
        return 'break'
    
    def focus_suggestion_up(self, event):
        """Xử lý phím Up trong search entry"""
        if self.suggestion_listbox.winfo_viewable() and self.suggestion_listbox.size() > 0:
            self.suggestion_listbox.focus_set()
            # Chọn item cuối cùng
            last_index = self.suggestion_listbox.size() - 1
            self.suggestion_listbox.activate(last_index)
            self.suggestion_listbox.selection_clear(0, tk.END)
            self.suggestion_listbox.selection_set(last_index)
        return 'break'
    
    def on_suggestion_up(self, event):
        """Xử lý phím Up trong suggestion listbox"""
        try:
            current = self.suggestion_listbox.curselection()
            if current:
                current_index = current[0]
                if current_index > 0:
                    new_index = current_index - 1
                    self.suggestion_listbox.selection_clear(0, tk.END)
                    self.suggestion_listbox.selection_set(new_index)
                    self.suggestion_listbox.activate(new_index)
                else:
                    # Quay về search entry
                    self.search_entry.focus_set()
                    self.hide_suggestions()
        except:
            pass
        return 'break'
    
    def on_suggestion_down(self, event):
        """Xử lý phím Down trong suggestion listbox"""
        try:
            current = self.suggestion_listbox.curselection()
            if current:
                current_index = current[0]
                if current_index < self.suggestion_listbox.size() - 1:
                    new_index = current_index + 1
                    self.suggestion_listbox.selection_clear(0, tk.END)
                    self.suggestion_listbox.selection_set(new_index)
                    self.suggestion_listbox.activate(new_index)
        except:
            pass
        return 'break'
    
    def get_word_info(self, word):
        """Lấy thông tin từ Cambridge - OPTIMIZED EXTREME"""
        url = f"https://dictionary.cambridge.org/dictionary/english/{word.lower()}"
        
        try:
            response = self.session.get(url, timeout=6)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'lxml')
            
            # Kiểm tra từ có tồn tại trong Cambridge Dictionary
            if self._is_word_not_found(soup):
                return None
            
            phonetics = self._get_phonetics_fast(soup)
            definitions = self._get_definitions_fast(soup)
            
            # Chỉ dịch nếu từ tồn tại - dùng Google Translate cho tra từ vựng thông thường
            if definitions:  # Chỉ dịch nếu có definitions
                word_meaning_vi = self._google_translate_text(word)
            else:
                word_meaning_vi = ""
            
            # Lấy audio URLs từ Cambridge
            audio_urls = self._get_audio_urls(soup)
            
            return {
                'word': word,
                'phonetic_uk': phonetics['uk'],
                'phonetic_us': phonetics['us'],
                'audio_uk': audio_urls['uk'],
                'audio_us': audio_urls['us'],
                'word_meaning_vi': word_meaning_vi,
                'definitions': definitions
            }
            
        except Exception as e:
            print(f"Error: {e}")
            return None
    
    def _is_word_not_found(self, soup):
        """Kiểm tra từ có tồn tại trong Cambridge Dictionary"""
        try:
            # Kiểm tra các dấu hiệu từ không tồn tại
            not_found_indicators = [
                "Sorry, we couldn't find any results",
                "No results found",
                "We couldn't find any results",
                "Sorry, no results found"
            ]
            
            page_text = soup.get_text().lower()
            for indicator in not_found_indicators:
                if indicator.lower() in page_text:
                    return True
            
            # Kiểm tra có definitions không
            definitions = self._get_definitions_fast(soup)
            return len(definitions) == 0
            
        except:
            return False
    
    def _translate_definitions_async(self, definitions):
        """Pre-translate các định nghĩa trong background"""
        def translate_all():
            for defn in definitions:
                # Pre-load vào cache
                self.translate_text(defn['definition'])
                for ex in defn['examples']:
                    self.translate_text(ex)
        
        # Chạy trong thread riêng, không block UI
        threading.Thread(target=translate_all, daemon=True).start()
    
    def _get_phonetics_fast(self, soup):
        """Lấy phiên âm UK và US"""
        result = {'uk': '', 'us': ''}
        
        try:
            pron_section = soup.find('div', class_='pos-header')
            if not pron_section:
                return result
            
            uk_div = pron_section.find('span', class_='uk')
            if uk_div:
                uk_ipa = uk_div.find('span', class_='ipa')
                if uk_ipa:
                    result['uk'] = uk_ipa.text.strip()
            
            us_div = pron_section.find('span', class_='us')
            if us_div:
                us_ipa = us_div.find('span', class_='ipa')
                if us_ipa:
                    result['us'] = us_ipa.text.strip()
                    
        except Exception as e:
            print(f"Error getting phonetics: {e}")
        
        return result
    
    def _get_audio_urls(self, soup):
        """Lấy URL audio từ Cambridge"""
        result = {'uk': '', 'us': ''}
        
        try:
            # Tìm audio UK
            uk_audio = soup.find('source', {'type': 'audio/mpeg', 'src': lambda x: x and 'uk' in x.lower()})
            if uk_audio and uk_audio.get('src'):
                result['uk'] = 'https://dictionary.cambridge.org' + uk_audio['src']
            
            # Tìm audio US
            us_audio = soup.find('source', {'type': 'audio/mpeg', 'src': lambda x: x and 'us' in x.lower()})
            if us_audio and us_audio.get('src'):
                result['us'] = 'https://dictionary.cambridge.org' + us_audio['src']
                
        except Exception as e:
            print(f"Error getting audio: {e}")
        
        return result
    
    def _get_definitions_fast(self, soup):
        """Lấy nghĩa - OPTIMIZED"""
        definitions = []
        
        try:
            entry_body = soup.find('div', class_='entry-body')
            if not entry_body:
                return [{'pos': '', 'definition': 'No definition found', 'examples': []}]
            
            pos_bodies = entry_body.find_all('div', class_='pos-body', limit=3)
            
            for pos_body in pos_bodies:
                pos_header = pos_body.find_previous_sibling('div', class_='pos-header')
                pos = ''
                if pos_header:
                    pos_elem = pos_header.find('span', class_='pos')
                    if pos_elem:
                        pos = pos_elem.text.strip()
                
                def_blocks = pos_body.find_all('div', class_='def-block', limit=2)
                
                for block in def_blocks:
                    def_elem = block.find('div', class_='def')
                    if def_elem:
                        definition = def_elem.text.strip()
                        
                        examples = []
                        example_elems = block.find_all('span', class_='eg', limit=2)
                        for ex in example_elems:
                            examples.append(ex.text.strip())
                        
                        definitions.append({
                            'pos': pos,
                            'definition': definition,
                            'examples': examples
                        })
                        pos = ''
            
        except Exception as e:
            print(f"Error: {e}")
        
        return definitions if definitions else [{'pos': '', 'definition': 'No definition found', 'examples': []}]
    
    def _show_error(self, word):
        self.clear_results()
        error_frame = tk.Frame(self.scrollable_frame, bg="white")
        error_frame.pack(pady=50)
        
        error_label = tk.Label(
            error_frame,
            text=f"Word '{word}' not found",
            font=("Arial", 16, "bold"),
            bg="white",
            fg="#E74C3C"
        )
        error_label.pack(pady=10)
        
        hint_label = tk.Label(
            error_frame,
            text="Please check the spelling and try again.",
            font=("Arial", 12),
            bg="white",
            fg="#666666"
        )
        hint_label.pack()
    
    def _display_results(self, word_info):
        """Hiển thị kết quả với giao diện đẹp"""
        self.clear_results()
        self.current_word_info = word_info

        content_frame = tk.Frame(self.scrollable_frame, bg=self.colors['white'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=20)
        
        # Word Title với nghĩa tiếng Việt
        title_frame = tk.Frame(content_frame, bg=self.colors['white'])
        title_frame.pack(anchor=tk.W, pady=(0, 15))
        
        word_title = tk.Label(
            title_frame, 
            text=word_info['word'].upper(), 
            font=("Segoe UI", 32, "bold"), 
            bg=self.colors['white'], 
            fg=self.colors['primary']
        )
        word_title.pack(side=tk.LEFT)
        
        # AI Vietnamese meaning
        self.ai_vi_label = tk.Label(
            title_frame, 
            text=(f"  •  {word_info['word_meaning_vi']}" if word_info.get('word_meaning_vi') else ""), 
            font=("Segoe UI", 18), 
            bg=self.colors['white'], 
            fg=self.colors['secondary']
        )
        self.ai_vi_label.pack(side=tk.LEFT, padx=(15, 0))
        
        # Pronunciation
        pron_frame = tk.Frame(content_frame, bg=self.colors['white'])
        pron_frame.pack(anchor=tk.W, pady=(0, 20))
        
        if word_info['phonetic_uk']:
            uk_frame = tk.Frame(pron_frame, bg=self.colors['white'])
            uk_frame.pack(side=tk.LEFT, padx=(0, 30))
            
            tk.Label(uk_frame, text="UK", font=("Segoe UI", 11, "bold"), bg=self.colors['white'], fg=self.colors['dark']).pack(side=tk.LEFT, padx=(0, 8))
            
            tk.Button(
                uk_frame,
                text="🔊",
                font=("Segoe UI", 18),
                bg=self.colors['white'],
                fg=self.colors['secondary'],
                relief=tk.FLAT,
                cursor="hand2",
                borderwidth=0,
                command=lambda: self.play_audio(word_info.get('audio_uk', ''), word_info['word'], 'uk')
            ).pack(side=tk.LEFT)
            
            uk_ipa_label = tk.Label(
                uk_frame,
                text=f"/{word_info['phonetic_uk']}/",
                font=("Segoe UI", 16),
                bg=self.colors['white'],
                fg=self.colors['danger']
            )
            uk_ipa_label.pack(side=tk.LEFT, padx=(8, 4))
            tk.Button(
                uk_frame,
                text="Copy",
                font=("Segoe UI", 9),
                bg=self.colors['light'],
                fg=self.colors['dark'],
                relief=tk.SOLID,
                borderwidth=1,
                command=lambda t=uk_ipa_label.cget('text'): self.copy_to_clipboard(t)
            ).pack(side=tk.LEFT)
        
        if word_info['phonetic_us']:
            us_frame = tk.Frame(pron_frame, bg=self.colors['white'])
            us_frame.pack(side=tk.LEFT)
            
            tk.Label(us_frame, text="US", font=("Segoe UI", 11, "bold"), bg=self.colors['white'], fg=self.colors['dark']).pack(side=tk.LEFT, padx=(0, 8))
            
            tk.Button(
                us_frame,
                text="🔊",
                font=("Segoe UI", 18),
                bg=self.colors['white'],
                fg=self.colors['secondary'],
                relief=tk.FLAT,
                cursor="hand2",
                borderwidth=0,
                command=lambda: self.play_audio(word_info.get('audio_us', ''), word_info['word'], 'us')
            ).pack(side=tk.LEFT)
            
            us_ipa_label = tk.Label(
                us_frame,
                text=f"/{word_info['phonetic_us']}/",
                font=("Segoe UI", 16),
                bg=self.colors['white'],
                fg=self.colors['danger']
            )
            us_ipa_label.pack(side=tk.LEFT, padx=(8, 4))
            tk.Button(
                us_frame,
                text="Copy",
                font=("Segoe UI", 9),
                bg=self.colors['light'],
                fg=self.colors['dark'],
                relief=tk.SOLID,
                borderwidth=1,
                command=lambda t=us_ipa_label.cget('text'): self.copy_to_clipboard(t)
            ).pack(side=tk.LEFT)
        
        # Separator
        tk.Frame(content_frame, height=2, bg=self.colors['border']).pack(fill=tk.X, pady=15)
        
        # Action buttons
        action_frame = tk.Frame(content_frame, bg=self.colors['white'])
        action_frame.pack(anchor=tk.W, pady=(0, 20))
        
        add_vocab_btn = tk.Button(
            action_frame,
            text="Add to Vocabulary",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['accent'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            command=self.add_current_word
        )
        add_vocab_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Definitions
        translation_targets = []
        for idx, definition in enumerate(word_info['definitions'], 1):
            self._display_definition(content_frame, idx, definition, translation_targets)
        if translation_targets:
            self._run_batch_translation(translation_targets)
        # AI translate chỉ khi user click nút "AI dịch theo ngữ cảnh"

    def copy_to_clipboard(self, text: str):
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()  # keep clipboard after window closed
        except Exception as e:
            print(f"Copy failed: {e}")
    
    def _display_definition(self, parent, idx, definition, translation_targets):
        """Hiển thị định nghĩa với giao diện đẹp"""
        def_container = tk.Frame(parent, bg=self.colors['white'])
        def_container.pack(fill=tk.X, pady=15, anchor=tk.W)
        
        # Thêm đường kẻ phân biệt giữa các nghĩa cùng mức độ
        if idx > 1:
            separator = tk.Frame(def_container, height=1, bg=self.colors['light_gray'])
            separator.pack(fill=tk.X, pady=(0, 15))
        
        # Part of speech
        if definition['pos']:
            pos_label = tk.Label(
                def_container,
                text=definition['pos'],
                font=("Segoe UI", 14, "bold"),
                bg=self.colors['white'],
                fg=self.colors['primary']
            )
            pos_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Definition row
        def_row = tk.Frame(def_container, bg=self.colors['white'])
        def_row.pack(fill=tk.X, pady=(0, 10))
        
        # Left: English
        left_frame = tk.Frame(def_row, bg=self.colors['white'])
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
        
        en_def_frame = tk.Frame(left_frame, bg=self.colors['white'])
        en_def_frame.pack(fill=tk.X)
        
        tk.Label(
            en_def_frame,
            text=f"{idx}.",
            font=("Segoe UI", 14, "bold"),
            bg=self.colors['white'],
            fg=self.colors['primary']
        ).pack(side=tk.LEFT, anchor=tk.N, padx=(0, 8))
        
        def_text_label = tk.Label(
            en_def_frame,
            text=definition['definition'],
            font=("Segoe UI", 14, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark'],
            wraplength=400,
            justify=tk.LEFT
        )
        def_text_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(
            en_def_frame,
            text="Copy",
            font=("Segoe UI", 9),
            bg=self.colors['light'],
            fg=self.colors['dark'],
            relief=tk.SOLID,
            borderwidth=1,
            command=lambda t=definition['definition']: self.copy_to_clipboard(t)
        ).pack(side=tk.LEFT, padx=(6, 0))
        
        # Right: Vietnamese - PLACEHOLDER trước, cập nhật sau
        right_frame = tk.Frame(def_row, bg=self.colors['light'], relief=tk.SOLID, borderwidth=1)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH)

        vi_label = tk.Label(
            right_frame,
            text="🇻🇳  Đang dịch...",
            font=("Segoe UI", 12),
            bg=self.colors['light'],
            fg=self.colors['secondary'],
            wraplength=300,
            justify=tk.LEFT,
            padx=12,
            pady=8
        )
        vi_label.pack()

        translation_targets.append({
            'text': definition['definition'],
            'label': vi_label,
            'prefix': "🇻🇳  "
        })
        
        # Examples
        if definition['examples']:
            for ex in definition['examples']:
                ex_frame = tk.Frame(left_frame, bg=self.colors['white'])
                ex_frame.pack(fill=tk.X, pady=(5, 0), padx=(25, 0))
                
                ex_text_label = tk.Label(
                    ex_frame,
                    text=f"• {ex}",
                    font=("Segoe UI", 13, "bold"),
                    bg=self.colors['white'],
                    fg=self.colors['dark'],
                    wraplength=400,
                    justify=tk.LEFT
                )
                ex_text_label.pack(side=tk.LEFT, anchor=tk.W)
                tk.Button(
                    ex_frame,
                    text="Copy",
                    font=("Segoe UI", 9),
                    bg=self.colors['light'],
                    fg=self.colors['dark'],
                    relief=tk.SOLID,
                    borderwidth=1,
                    command=lambda t=ex: self.copy_to_clipboard(t)
                ).pack(side=tk.LEFT, padx=(6,0))
                
                # Vietnamese example - placeholder
                vi_ex_label = tk.Label(
                    ex_frame,
                    text="  ...",
                    font=("Segoe UI", 11),
                    bg=self.colors['white'],
                    fg=self.colors['secondary'],
                    wraplength=400,
                    justify=tk.LEFT
                )
                vi_ex_label.pack(anchor=tk.W, pady=(2, 0))

                translation_targets.append({
                    'text': ex,
                    'label': vi_ex_label,
                    'prefix': "  "
                })

    def run_ai_translate_current(self):
        if not self.gemini_enabled:
            return
        word_info = getattr(self, 'current_word_info', None)
        if not word_info:
            return
        word = word_info['word']
        context_widget = getattr(self, 'context_text', None)
        context = context_widget.get('1.0', tk.END).strip() if context_widget else ""
        defs_en = [d.get('definition', '') for d in word_info.get('definitions', [])][:3]

        # UI: show loading and disable button
        try:
            if hasattr(self, 'ai_status_var'):
                self.ai_status_var.set("Đang gọi AI…")
            if hasattr(self, 'ai_translate_btn'):
                self.ai_translate_btn.configure(
                    state=tk.DISABLED,
                    text="Đang dịch…",
                    bg=self.colors['border'],
                    fg=self.colors['dark']
                )
        except Exception:
            pass

        def worker():
            vi, example_en = self.ai_translate_with_context(word, defs_en, context)
            def done():
                try:
                    if vi and getattr(self, 'ai_vi_label', None):
                        self.ai_vi_label.config(text=f"  •  {vi}")
                    if hasattr(self, 'ai_status_var'):
                        if vi:
                            self.ai_status_var.set("Hoàn tất")
                        else:
                            self.ai_status_var.set("AI không trả kết quả - Kiểm tra console")
                    if hasattr(self, 'ai_translate_btn'):
                        self.ai_translate_btn.configure(
                            state=tk.NORMAL,
                            text="AI dịch theo ngữ cảnh",
                            bg=self.colors['secondary'] if self.gemini_enabled else self.colors['border'],
                            fg=self.colors['white'] if self.gemini_enabled else self.colors['dark']
                        )
                except Exception:
                    pass
                self.current_ai_vi = vi
                self.current_ai_example_en = example_en
            self.root.after(0, done)
        threading.Thread(target=worker, daemon=True).start()

    def ai_translate_with_context(self, word: str, defs_en: list, context: str):
        max_retries = len(self.api_keys)
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if not self.gemini_enabled:
                    return None, None
                    
                prompt = (
                    "You are a bilingual English-Vietnamese lexicographer. Given an English headword, its brief glosses, "
                    "and an optional user-provided context, produce: (1) a concise Vietnamese meaning of the headword "
                    "that best fits the context (≤6 words), and (2) one short English example sentence that naturally uses the word. "
                    "Output as JSON: {\"vi_meaning\": string, \"example_en\": string}.\n\n"
                    f"Headword: {word}\nGlosses: {defs_en}\nContext: {context or '(none)'}\n"
                )
                print(f"[AI] Calling Gemini for '{word}' with context: {context[:50] if context else '(none)'}")
                resp = self.gemini_model.generate_content(prompt)
                text = (getattr(resp, 'text', None) or '').strip()
                print(f"[AI] Response: {text[:200]}")
                vi, ex = None, None
                import json as _json, re as _re
                m = _re.search(r"\{[\s\S]*\}", text)
                if m:
                    try:
                        obj = _json.loads(m.group(0))
                        vi = obj.get('vi_meaning')
                        ex = obj.get('example_en')
                        print(f"[AI] Parsed: vi={vi}, ex={ex}")
                    except Exception as parse_err:
                        print(f"[AI] JSON parse error: {parse_err}")
                if not vi and text:
                    vi = text.split('\n')[0][:60]
                    print(f"[AI] Fallback vi from first line: {vi}")
                if not ex:
                    ex = f"{word.capitalize()} is used naturally in this context."
                return (vi.strip() if vi else None), (ex.strip() if ex else None)
                
            except Exception as e:
                error_msg = str(e)
                print(f"[AI] Error: {type(e).__name__}: {error_msg}")
                
                # Handle API error and try switching keys
                error_type = self._handle_api_error(error_msg)
                
                if error_type == "switched":
                    retry_count += 1
                    print(f"[AI] Retrying with new key (attempt {retry_count}/{max_retries})")
                    continue  # Retry with new key
                elif error_type == "exhausted":
                    status_msg = "Tất cả API keys đã hết quota"
                elif "api" in error_msg.lower() and "key" in error_msg.lower():
                    status_msg = "API key không hợp lệ"
                elif "network" in error_msg.lower() or "connection" in error_msg.lower():
                    status_msg = "Lỗi kết nối mạng"
                else:
                    status_msg = f"Lỗi: {type(e).__name__}"
                
                try:
                    if hasattr(self, 'ai_status_var'):
                        self.ai_status_var.set(status_msg)
                    if hasattr(self, 'ai_translate_btn'):
                        self.ai_translate_btn.configure(
                            state=tk.NORMAL,
                            text="AI dịch theo ngữ cảnh",
                            bg=self.colors['secondary'] if self.gemini_enabled else self.colors['border'],
                            fg=self.colors['white'] if self.gemini_enabled else self.colors['dark']
                        )
                except Exception:
                    pass
                return None, None
        
        # If all retries exhausted
        return None, None
    
    def _initialize_gemini(self):
        """Initialize Gemini with current API key"""
        if self.current_key_index < len(self.api_keys):
            try:
                genai.configure(api_key=self.api_keys[self.current_key_index])
                self.gemini_model = genai.GenerativeModel('gemini-2.0-flash')
                self.gemini_enabled = True
                print(f"[Gemini] ✅ Initialized with API key #{self.current_key_index + 1}/{len(self.api_keys)}")
                return True
            except Exception as e:
                print(f"[Gemini] ❌ Key #{self.current_key_index + 1} failed: {e}")
                return False
        else:
            print("[Gemini] ⚠️ No more API keys available")
            self.gemini_enabled = False
            self.gemini_model = None
            return False
    
    def _switch_to_next_key(self):
        """Switch to next available API key when quota exceeded"""
        old_index = self.current_key_index
        self.current_key_index += 1
        
        if self.current_key_index < len(self.api_keys):
            print(f"[Gemini] 🔄 Switching from key #{old_index + 1} to key #{self.current_key_index + 1}")
            success = self._initialize_gemini()
            if success:
                # Show toast notification
                self.root.after(0, lambda: self.show_toast(f"Chuyển sang API key #{self.current_key_index + 1}"))
                return True
            else:
                # Try next key recursively
                return self._switch_to_next_key()
        else:
            print("[Gemini] 🚫 All API keys exhausted - AI features disabled")
            self.gemini_enabled = False
            self.gemini_model = None
            self.root.after(0, lambda: self.show_toast("Tất cả API keys đã hết quota"))
            return False
    
    def _handle_api_error(self, error_msg):
        """Handle API errors and switch keys if quota exceeded"""
        if "quota" in error_msg.lower() or "429" in error_msg:
            print(f"[Gemini] ⚠️ Quota exceeded for key #{self.current_key_index + 1}")
            if self._switch_to_next_key():
                return "switched"  # Successfully switched to new key
            else:
                return "exhausted"  # All keys exhausted
        return "other"  # Other error types
    
    def play_audio(self, audio_url, word, region='uk'):
        """Phát audio từ Cambridge hoặc fallback sang TTS"""
        if audio_url:
            threading.Thread(target=self._play_cambridge_audio, args=(audio_url,), daemon=True).start()
        else:
            # Fallback: dùng TTS
            if self.tts_engine:
                threading.Thread(target=self._pronounce_thread, args=(word,), daemon=True).start()
    
    def _play_cambridge_audio(self, audio_url):
        """Phát audio Cambridge trực tiếp trong app với pygame"""
        try:
            import pygame

            print(f"🔊 Playing Cambridge audio: {audio_url}")

            cache_path = None
            if self.audio_cache_dir:
                filename = hashlib.sha256(audio_url.encode('utf-8')).hexdigest() + '.mp3'
                cache_path = os.path.join(self.audio_cache_dir, filename)

            audio_source = None
            if cache_path and os.path.exists(cache_path):
                audio_source = cache_path
            else:
                response = self.session.get(audio_url, timeout=5)
                if response.status_code == 200:
                    if cache_path:
                        try:
                            with open(cache_path, 'wb') as f:
                                f.write(response.content)
                            audio_source = cache_path
                        except Exception as e:
                            print(f"[Audio Cache] Failed to write cache: {e}")
                    if not audio_source:
                        from io import BytesIO
                        audio_source = BytesIO(response.content)
                else:
                    print(f"❌ Cambridge audio failed: {response.status_code}")

            if not audio_source:
                # Fallback sang TTS
                if self.tts_engine and hasattr(self, 'current_word_info') and self.current_word_info:
                    threading.Thread(target=self._pronounce_thread, args=(self.current_word_info['word'],), daemon=True).start()
                return

            if not pygame.mixer.get_init():
                pygame.mixer.init()

            if hasattr(audio_source, 'seek'):
                audio_source.seek(0)

            pygame.mixer.music.load(audio_source)
            pygame.mixer.music.play()

            # Đợi phát xong
            while pygame.mixer.music.get_busy():
                time.sleep(0.1)

        except Exception as e:
            print(f"Error playing Cambridge audio: {e}")
            # Fallback sang TTS nếu lỗi
            if self.tts_engine and hasattr(self, 'current_word_info') and self.current_word_info:
                threading.Thread(target=self._pronounce_thread, args=(self.current_word_info['word'],), daemon=True).start()
    
    def _pronounce_thread(self, word):
        """Fallback TTS"""
        try:
            if self.tts_engine:
                self.tts_engine.say(word)
                self.tts_engine.runAndWait()
        except Exception as e:
            print(f"Error: {e}")
    
    def add_current_word(self):
        if hasattr(self, 'current_word_info'):
            self.add_to_vocabulary(
                self.current_word_info['word'],
                self.current_word_info['phonetic_uk'],
                self.current_word_info['phonetic_us'],
                self.current_word_info['definitions'],
                getattr(self, 'current_ai_vi', None),
                getattr(self, 'current_ai_example_en', None)
            )
        else:
            self.show_toast("Vui lòng tra từ trước!")
    
    def export_to_excel(self, clear_after_export=False):
        """Export từ vựng ra Excel cho Quizlet"""
        if not self.vocabulary:
            self.show_toast("Chưa có từ vựng nào để export!")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from tkinter import filedialog
            
            # Hộp chọn tuỳ chọn export
            options = tk.Toplevel(self.root)
            options.title("Tùy chọn Export")
            options.geometry("380x180")
            del_var = tk.BooleanVar(value=True)
            tk.Label(options, text="Tùy chọn trước khi export:", font=("Arial", 12, "bold")).pack(pady=10)
            tk.Checkbutton(options, text="Xóa TẤT CẢ từ vựng sau khi export", variable=del_var).pack()
            confirmed = {'ok': False}
            def _ok():
                confirmed['ok'] = True
                options.destroy()
            def _cancel():
                options.destroy()
            btnf = tk.Frame(options)
            btnf.pack(pady=15)
            tk.Button(btnf, text="OK", width=10, command=_ok).pack(side=tk.LEFT, padx=6)
            tk.Button(btnf, text="Hủy", width=10, command=_cancel).pack(side=tk.LEFT)
            options.transient(self.root)
            options.grab_set()
            self.root.wait_window(options)
            if not confirmed['ok']:
                return
            clear_after_export = del_var.get()
            
            # Chọn nơi lưu file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="my_vocabulary_quizlet.xlsx"
            )
            
            if not file_path:
                return
            
            # Tạo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Vocabulary"
            
            # Header
            ws['A1'] = "Từ vựng / Cụm từ (Exam dưới)"
            ws['B1'] = "Nghĩa"
            
            # Style header
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            ws['A1'].font = header_font
            ws['B1'].font = header_font
            ws['A1'].fill = header_fill
            ws['B1'].fill = header_fill
            
            # Điều chỉnh độ rộng cột
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 40
            
            # Thêm dữ liệu
            row = 2
            for item in self.vocabulary:
                word = item['word']
                phonetic_uk = item.get('phonetic_uk', '')
                phonetic_us = item.get('phonetic_us', '')
                definitions = item.get('definitions', [])
                
                # Lấy ví dụ: ưu tiên AI example nếu có
                example = item.get('ai_example_en') or ""
                if not example and definitions and definitions[0].get('examples'):
                    example = definitions[0]['examples'][0]
                
                # Cột A: Từ + IPA + Example
                cell_a_text = f"{word}"
                if phonetic_uk or phonetic_us:
                    ipa = phonetic_uk if phonetic_uk else phonetic_us
                    cell_a_text += f" /{ipa}/"
                if example:
                    cell_a_text += f"\nExam: {example}"
                
                # Cột B: Nghĩa tiếng Việt - ưu tiên ai_vi nếu có
                meaning_vi = item.get('ai_vi') or self.translate_text(word)
                # Thêm dấu chấm phẩy ở cuối nghĩa tiếng Việt
                if meaning_vi and not meaning_vi.endswith(';'):
                    meaning_vi += ';'
                
                ws[f'A{row}'] = cell_a_text
                ws[f'B{row}'] = meaning_vi
                
                # Style cells
                ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
            
            # Lưu file
            wb.save(file_path)
            
            # Thông báo thành công
            word_count = len(self.vocabulary)
            self.show_toast(f"Đã export {word_count} từ vào file!")
            
            # Xóa từ vựng nếu user chọn
            if clear_after_export:
                self.vocabulary = []
                self.save_vocabulary()
                self.show_toast(f"Đã xóa {word_count} từ vựng cũ!")
            
        except Exception as e:
            self.show_toast(f"Lỗi: {str(e)}")
    
    def show_vocabulary(self):
        vocab_window = tk.Toplevel(self.root)
        vocab_window.title("Danh sách từ vựng")
        vocab_window.geometry("700x500")
        
        self.vocab_title_label = tk.Label(
            vocab_window,
            text=f"Từ vựng của tôi ({len(self.vocabulary)} từ)",
            font=("Arial", 16, "bold"),
            bg="#8E44AD",
            fg="white",
            pady=15
        )
        self.vocab_title_label.pack(fill=tk.X)
        
        list_frame = tk.Frame(vocab_window)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        vocab_listbox = tk.Listbox(list_frame, font=("Arial", 12), yscrollcommand=scrollbar.set)
        vocab_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=vocab_listbox.yview)
        
        for item in self.vocabulary:
            vocab_listbox.insert(tk.END, f"{item['word']}  -  {item['added_date']}")
        
        btn_frame = tk.Frame(vocab_window)
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def view_word():
            selection = vocab_listbox.curselection()
            if selection:
                word = self.vocabulary[selection[0]]['word']
                self.search_entry.delete(0, tk.END)
                self.search_entry.insert(0, word)
                vocab_window.destroy()
                self.search_word()
        
        def delete_word():
            selection = vocab_listbox.curselection()
            if selection:
                idx = selection[0]
                word = self.vocabulary[idx]['word']
                if messagebox.askyesno("Xác nhận", f"Xóa '{word}'?"):
                    del self.vocabulary[idx]
                    self.save_vocabulary()
                    vocab_listbox.delete(idx)
                    if getattr(self, "vocab_title_label", None):
                        self.vocab_title_label.config(
                            text=f"📚 Từ vựng của tôi ({len(self.vocabulary)} từ)"
                        )
        
        def delete_all():
            if not self.vocabulary:
                self.show_toast("Danh sách đã trống!")
                return
            
            count = len(self.vocabulary)
            if messagebox.askyesno(
                "⚠️ Xác nhận xóa TẤT CẢ", 
                f"Bạn chắc chắn muốn xóa TẤT CẢ {count} từ?\n\n"
                "Hành động này KHÔNG THỂ hoàn tác!"
            ):
                self.vocabulary = []
                self.save_vocabulary()
                vocab_listbox.delete(0, tk.END)
                self.show_toast(f"Đã xóa {count} từ!")
                vocab_window.destroy()
        
        tk.Button(
            btn_frame, 
            text="Xem từ", 
            font=("Arial", 11), 
            bg="#3498DB", 
            fg="white", 
            padx=20, 
            pady=8, 
            command=view_word
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(
            btn_frame, 
            text="Xóa 1 từ", 
            font=("Arial", 11), 
            bg="#E74C3C", 
            fg="white", 
            padx=15, 
            pady=8, 
            command=delete_word
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(
            btn_frame, 
            text="Xóa TẤT CẢ", 
            font=("Arial", 11, "bold"), 
            bg="#C0392B", 
            fg="white", 
            padx=15, 
            pady=8, 
            command=delete_all
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(
            btn_frame, 
            text="Export Excel", 
            font=("Arial", 11, "bold"), 
            bg="#27AE60", 
            fg="white", 
            padx=20, 
            pady=8, 
            command=self.export_to_excel
        ).pack(side=tk.LEFT)

    def show_text_translator(self):
        """Hiển thị giao diện dịch văn bản AI - như phiên bản cũ"""
        # Clear current content
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.update_current_word_display("Dịch văn bản AI")
        
        # Main translator container
        translator_frame = tk.Frame(self.scrollable_frame, bg=self.colors['white'])
        translator_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Title
        title_label = tk.Label(
            translator_frame,
            text="Dịch văn bản AI",
            font=("Segoe UI", 24, "bold"),
            bg=self.colors['white'],
            fg=self.colors['primary']
        )
        title_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Description
        desc_label = tk.Label(
            translator_frame,
            text="Paste văn bản tiếng Anh vào ô bên dưới để dịch sang tiếng Việt (hỗ trợ dịch theo ngữ cảnh)",
            font=("Segoe UI", 11),
            bg=self.colors['white'],
            fg="#64748b"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 20))
        
        # Main translation area - Google Translate style
        main_frame = tk.Frame(translator_frame, bg=self.colors['white'])
        main_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Left column - Input
        left_column = tk.Frame(main_frame, bg=self.colors['white'])
        left_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Input label
        tk.Label(
            left_column,
            text="Văn bản cần dịch:",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(anchor=tk.W, pady=(0, 8))
        
        # Input text area with auto-resize
        self.translate_input = tk.Text(
            left_column,
            height=8,
            font=("Segoe UI", 11),
            wrap=tk.WORD,
            bg=self.colors['light'],
            relief=tk.SOLID,
            borderwidth=1,
            padx=15,
            pady=12
        )
        self.translate_input.pack(fill=tk.BOTH, expand=True)
        
        # Auto-resize only
        self.translate_input.bind('<KeyPress>', self._on_input_change)
        self.translate_input.bind('<KeyRelease>', self._on_input_change)
        
        # Right column - Translation
        right_column = tk.Frame(main_frame, bg=self.colors['white'])
        right_column.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        # Translation label
        tk.Label(
            right_column,
            text="Bản dịch:",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(anchor=tk.W, pady=(0, 8))
        
        # Translation text area
        self.translate_output = tk.Text(
            right_column,
            height=8,
            font=("Segoe UI", 11),
            wrap=tk.WORD,
            bg="#f0fdf4",
            relief=tk.SOLID,
            borderwidth=1,
            padx=15,
            pady=12,
            state=tk.DISABLED
        )
        self.translate_output.pack(fill=tk.BOTH, expand=True)
        
        # Button frame
        btn_frame = tk.Frame(translator_frame, bg=self.colors['white'])
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Translate button
        translate_btn = tk.Button(
            btn_frame,
            text="Dịch",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=30,
            pady=8,
            cursor="hand2",
            command=self.translate_text_ui
        )
        translate_btn.pack(side=tk.LEFT)
        
        # Clear button
        clear_btn = tk.Button(
            btn_frame,
            text="Xóa",
            font=("Segoe UI", 11),
            bg="#ef4444",
            fg=self.colors['white'],
            relief=tk.FLAT,
            padx=20,
            pady=8,
            cursor="hand2",
            command=lambda: self._clear_translator_ui()
        )
        clear_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # Vocabulary explanation section
        vocab_section = tk.Frame(translator_frame, bg=self.colors['white'])
        vocab_section.pack(fill=tk.X, pady=(15, 0))
        
        # Vocabulary explanation label
        tk.Label(
            vocab_section,
            text="Giải thích từ vựng:",
            font=("Segoe UI", 12, "bold"),
            bg=self.colors['white'],
            fg=self.colors['dark']
        ).pack(anchor=tk.W, pady=(0, 8))
        
        # Vocabulary explanation text area
        self.vocab_explanation = tk.Text(
            vocab_section,
            height=6,
            font=("Segoe UI", 10),
            wrap=tk.WORD,
            bg="#fef3c7",
            relief=tk.SOLID,
            borderwidth=1,
            padx=15,
            pady=12,
            state=tk.DISABLED
        )
        self.vocab_explanation.pack(fill=tk.X)
        
        # AI status
        if not self.gemini_enabled:
            warning_label = tk.Label(
                translator_frame,
                text="AI chưa khả dụng. Sẽ sử dụng Google Translate.",
                font=("Segoe UI", 10),
                bg=self.colors['white'],
                fg=self.colors['warning']
            )
            warning_label.pack(anchor=tk.W, pady=(10, 0))

    def translate_text_ui(self):
        """Dịch văn bản bằng AI từ UI"""
        text = self.translate_input.get("1.0", tk.END).strip()
        
        if not text:
            # Clear outputs if empty
            self.translate_output.config(state=tk.NORMAL)
            self.translate_output.delete("1.0", tk.END)
            self.translate_output.config(state=tk.DISABLED)
            self.vocab_explanation.config(state=tk.NORMAL)
            self.vocab_explanation.delete("1.0", tk.END)
            self.vocab_explanation.config(state=tk.DISABLED)
            return
        
        # Clear outputs
        self.translate_output.config(state=tk.NORMAL)
        self.translate_output.delete("1.0", tk.END)
        self.translate_output.insert("1.0", "🔄 Đang dịch...\n")
        self.translate_output.config(state=tk.DISABLED)
        
        self.vocab_explanation.config(state=tk.NORMAL)
        self.vocab_explanation.delete("1.0", tk.END)
        self.vocab_explanation.insert("1.0", "🔄 Đang phân tích từ vựng...\n")
        self.vocab_explanation.config(state=tk.DISABLED)
        self.root.update()
        
        # Get context from main context box (shared)
        context = ""
        if hasattr(self, 'context_text'):
            context = self.context_text.get("1.0", tk.END).strip()
        
        # Run translation in background thread
        threading.Thread(target=self._translate_text_thread, args=(text, context), daemon=True).start()
    
    def _translate_text_thread(self, text, context=""):
        """Thread dịch văn bản"""
        try:
            # Count words
            words = text.split()
            word_count = len(words)
            
            # If single word, suggest using dictionary search
            if word_count == 1:
                translation = self._translate_simple(text)
                vocab_explanation = self._get_vocab_explanation(text)
                self.root.after(0, lambda: self._update_translation_result(
                    f"Gợi ý: Đây là 1 từ đơn. Bạn có thể tra từ điển để xem nghĩa chi tiết hơn.\n\n"
                    f"Nghĩa: {translation}",
                    vocab_explanation
                ))
                return
            
            # For 2+ words, translate with context using AI (combined method to save API requests)
            if self.gemini_enabled:
                # Combined translation and vocabulary explanation in 1 AI request
                translation_result, vocab_explanation = self._translate_with_gemini_combined(text, context)
                # Final update with both results
                self.root.after(0, lambda: self._update_translation_result(translation_result, vocab_explanation))
            else:
                translation_result = self._translate_simple(text)
                vocab_explanation = "AI chưa khả dụng. Không thể phân tích từ vựng."
                self.root.after(0, lambda: self._update_translation_result(translation_result, vocab_explanation))
            
        except Exception as e:
            error_msg = f"Lỗi dịch: {str(e)}"
            print(error_msg)
            self.root.after(0, lambda: self._update_translation_result(error_msg, "Không thể phân tích từ vựng"))
    
    def _translate_with_gemini_combined(self, text, context=""):
        """Dịch văn bản và giải thích từ vựng trong 1 AI request để tiết kiệm API"""
        max_retries = len(self.api_keys)
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if not self.gemini_enabled:
                    translation = self._translate_simple(text)
                    vocab_explanation = "AI chưa khả dụng. Không thể phân tích từ vựng."
                    return translation, vocab_explanation
                    
                if context:
                    prompt = f"""Dịch văn bản tiếng Anh sau sang tiếng Việt và giải thích từ vựng quan trọng:

Văn bản: "{text}"
Ngữ cảnh: "{context}"

Yêu cầu:
1. Dịch tự nhiên, dễ hiểu, phù hợp ngữ cảnh
2. Giải thích CHỈ các từ/cụm từ quan trọng (bỏ qua: the, of, a, an, and, or, but, in, on, at, to, for, with, by, from, is, are, was, were, be, been, have, has, had, do, does, did, will, would, could, should, may, might, can, must, shall)
3. Format chính xác:
   BẢN DỊCH:
   [bản dịch tiếng Việt]
   
   TỪ VỰNG QUAN TRỌNG:
   • [từ 1]: [nghĩa tiếng Việt]
   • [từ 2]: [nghĩa tiếng Việt]
   • [từ 3]: [nghĩa tiếng Việt]"""
                else:
                    prompt = f"""Dịch văn bản tiếng Anh sau sang tiếng Việt và giải thích từ vựng quan trọng:

"{text}"

Yêu cầu:
1. Dịch tự nhiên, dễ hiểu
2. Giải thích CHỈ các từ/cụm từ quan trọng (bỏ qua: the, of, a, an, and, or, but, in, on, at, to, for, with, by, from, is, are, was, were, be, been, have, has, had, do, does, did, will, would, could, should, may, might, can, must, shall)
3. Format chính xác:
   BẢN DỊCH:
   [bản dịch tiếng Việt]
   
   TỪ VỰNG QUAN TRỌNG:
   • [từ 1]: [nghĩa tiếng Việt]
   • [từ 2]: [nghĩa tiếng Việt]
   • [từ 3]: [nghĩa tiếng Việt]"""

                print(f"[AI Combined] Translating and analyzing: {text[:50]}...")
                response = self.gemini_model.generate_content(prompt, stream=True)
                
                # Collect streaming text
                full_response = ""
                for chunk in response:
                    if hasattr(chunk, 'text') and chunk.text:
                        full_response += chunk.text
                        # Update UI in real-time with translation only
                        self.root.after(0, lambda t=full_response: self._update_translation_streaming_combined(t))
                
                # Parse response to separate translation and vocabulary
                translation, vocab_explanation = self._parse_combined_response(full_response.strip())
                return translation, vocab_explanation
                
            except Exception as e:
                error_msg = str(e)
                print(f"Gemini combined translation failed: {error_msg}")
                
                # Handle API error and try switching keys
                error_type = self._handle_api_error(error_msg)
                
                if error_type == "switched":
                    retry_count += 1
                    print(f"[Combined Translation] Retrying with new key (attempt {retry_count}/{max_retries})")
                    continue  # Retry with new key
                elif error_type == "exhausted":
                    error_display = "⚠️ Tất cả API keys đã hết - Sử dụng Google Translate"
                elif "api" in error_msg.lower() and "key" in error_msg.lower():
                    error_display = "⚠️ API key không hợp lệ - Sử dụng Google Translate"
                elif "network" in error_msg.lower() or "connection" in error_msg.lower():
                    error_display = "⚠️ Lỗi kết nối - Sử dụng Google Translate"
                else:
                    error_display = f"⚠️ Lỗi AI: {type(e).__name__} - Sử dụng Google Translate"
                
                # Cập nhật UI với thông báo lỗi
                self.root.after(0, lambda: self._update_translation_streaming_combined(error_display))
                
                # Fallback to Google Translate
                translation = self._translate_simple(text)
                vocab_explanation = "⚠️ Không thể phân tích từ vựng"
                return translation, vocab_explanation
        
        # If all retries exhausted, fallback
        translation = self._translate_simple(text)
        vocab_explanation = "⚠️ Không thể phân tích từ vựng"
        return translation, vocab_explanation
    
    def _parse_combined_response(self, response):
        """Parse response để tách translation và vocabulary explanation"""
        try:
            # Tìm phần BẢN DỊCH
            translation_start = response.find("BẢN DỊCH:")
            vocab_start = response.find("TỪ VỰNG QUAN TRỌNG:")
            
            translation = ""
            vocab_explanation = ""
            
            if translation_start != -1:
                if vocab_start != -1:
                    # Có cả 2 phần
                    translation = response[translation_start + 10:vocab_start].strip()
                    vocab_explanation = response[vocab_start:].strip()
                else:
                    # Chỉ có translation
                    translation = response[translation_start + 10:].strip()
                    vocab_explanation = "Không có từ vựng quan trọng"
            else:
                # Fallback: lấy dòng đầu làm translation
                lines = response.split('\n')
                translation = lines[0].strip() if lines else response
                vocab_explanation = "Không thể phân tích từ vựng"
            
            # Clean up translation
            if translation.startswith('"') and translation.endswith('"'):
                translation = translation[1:-1]
            if translation.startswith("'") and translation.endswith("'"):
                translation = translation[1:-1]
            
            # Clean up vocabulary
            if vocab_explanation and not vocab_explanation.startswith("TỪ VỰNG QUAN TRỌNG:"):
                vocab_explanation = f"TỪ VỰNG QUAN TRỌNG:\n\n{vocab_explanation}"
            
            return translation, vocab_explanation
            
        except Exception as e:
            print(f"[Parse Error] {e}")
            return response, "Không thể phân tích từ vựng"
    
    def _update_translation_streaming_combined(self, full_response):
        """Cập nhật UI với response đầy đủ (chỉ hiển thị translation)"""
        try:
            # Parse để lấy chỉ translation
            translation, _ = self._parse_combined_response(full_response)
            self.translate_output.config(state=tk.NORMAL)
            self.translate_output.delete("1.0", tk.END)
            self.translate_output.insert("1.0", translation)
            self.translate_output.config(state=tk.DISABLED)
            self.root.update()
        except Exception as e:
            print(f"[Streaming Error] {e}")
            self.translate_output.config(state=tk.NORMAL)
            self.translate_output.delete("1.0", tk.END)
            self.translate_output.insert("1.0", full_response)
            self.translate_output.config(state=tk.DISABLED)
            self.root.update()

    def _translate_with_gemini(self, text, context=""):
        """Dịch văn bản bằng Gemini AI với ngữ cảnh (legacy method)"""
        max_retries = len(self.api_keys)
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if not self.gemini_enabled:
                    return self._translate_simple(text)
                    
                if context:
                    prompt = f"""Dịch văn bản tiếng Anh sau sang tiếng Việt tự nhiên, sát nghĩa và phù hợp ngữ cảnh:

Văn bản: "{text}"
Ngữ cảnh: "{context}"

Yêu cầu:
- Dịch tự nhiên, dễ hiểu
- Giữ nguyên ý nghĩa
- Phù hợp với ngữ cảnh đã cho
- CHỈ trả về bản dịch tiếng Việt, KHÔNG giải thích thêm"""
                else:
                    prompt = f"""Dịch văn bản tiếng Anh sau sang tiếng Việt tự nhiên, sát nghĩa và phù hợp ngữ cảnh:

"{text}"

Yêu cầu:
- Dịch tự nhiên, dễ hiểu
- Giữ nguyên ý nghĩa
- Phù hợp ngữ cảnh
- CHỈ trả về bản dịch tiếng Việt, KHÔNG giải thích thêm"""

                # Streaming response
                response = self.gemini_model.generate_content(prompt, stream=True)
                
                # Collect streaming text
                translation_parts = []
                for chunk in response:
                    if hasattr(chunk, 'text') and chunk.text:
                        translation_parts.append(chunk.text)
                        # Update UI in real-time
                        current_text = ''.join(translation_parts)
                        self.root.after(0, lambda t=current_text: self._update_translation_streaming(t))
                
                translation = ''.join(translation_parts).strip()
                return translation
                
            except Exception as e:
                error_msg = str(e)
                print(f"Gemini translation failed: {error_msg}")
                
                # Handle API error and try switching keys
                error_type = self._handle_api_error(error_msg)
                
                if error_type == "switched":
                    retry_count += 1
                    print(f"[Translation] Retrying with new key (attempt {retry_count}/{max_retries})")
                    continue  # Retry with new key
                elif error_type == "exhausted":
                    error_display = "⚠️ Tất cả API keys đã hết - Sử dụng Google Translate"
                elif "api" in error_msg.lower() and "key" in error_msg.lower():
                    error_display = "⚠️ API key không hợp lệ - Sử dụng Google Translate"
                elif "network" in error_msg.lower() or "connection" in error_msg.lower():
                    error_display = "⚠️ Lỗi kết nối - Sử dụng Google Translate"
                else:
                    error_display = f"⚠️ Lỗi AI: {type(e).__name__} - Sử dụng Google Translate"
                
                # Cập nhật UI với thông báo lỗi
                self.root.after(0, lambda: self._update_translation_streaming(error_display))
                
                # Fallback to Google Translate
                return self._translate_simple(text)
        
        # If all retries exhausted, fallback
        return self._translate_simple(text)
    
    def _translate_simple(self, text):
        """Dịch văn bản bằng Google Translate"""
        try:
            translator = GoogleTranslator(source='en', target='vi')
            translation = translator.translate(text)
            return f"Google Translate:\n\n{translation}"
        except Exception as e:
            return f"Lỗi dịch: {str(e)}"
    
    def _update_translation_result(self, result, vocab_explanation=""):
        """Cập nhật kết quả dịch vào UI"""
        self.translate_output.config(state=tk.NORMAL)
        self.translate_output.delete("1.0", tk.END)
        self.translate_output.insert("1.0", result)
        self.translate_output.config(state=tk.DISABLED)
        
        # Update vocabulary explanation
        self.vocab_explanation.config(state=tk.NORMAL)
        self.vocab_explanation.delete("1.0", tk.END)
        self.vocab_explanation.insert("1.0", vocab_explanation)
        self.vocab_explanation.config(state=tk.DISABLED)
    
    def _update_translation_streaming(self, current_text):
        """Cập nhật kết quả dịch streaming (như ChatGPT)"""
        self.translate_output.config(state=tk.NORMAL)
        self.translate_output.delete("1.0", tk.END)
        self.translate_output.insert("1.0", current_text)
        self.translate_output.config(state=tk.DISABLED)
        self.root.update()
    
    def _get_vocab_explanation(self, text):
        """Lấy giải thích từ vựng chi tiết bằng AI"""
        max_retries = len(self.api_keys)
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                if not self.gemini_enabled:
                    return "AI chưa khả dụng. Không thể phân tích từ vựng."
                
                # Tách từ và cụm từ quan trọng
                words = text.split()
                if len(words) == 1:
                    # Từ đơn
                    prompt = f"""Phân tích từ tiếng Anh "{text}" và giải thích chi tiết:

1. Nghĩa chính (tiếng Việt)
2. Từ loại (danh từ, động từ, tính từ...)
3. Cách phát âm (nếu có)
4. Ví dụ sử dụng ngắn gọn
5. Từ đồng nghĩa (nếu có)

Trả lời ngắn gọn, dễ hiểu:"""
                else:
                    # Câu/đoạn văn - CHỈ giải thích từ quan trọng
                    prompt = f"""Phân tích văn bản tiếng Anh sau và giải thích CHỈ các từ/cụm từ quan trọng:

"{text}"

Yêu cầu:
- BỎ QUA các từ cơ bản: the, of, a, an, and, or, but, in, on, at, to, for, with, by, from, is, are, was, were, be, been, have, has, had, do, does, did, will, would, could, should, may, might, can, must, shall
- CHỈ giải thích từ có nghĩa quan trọng, từ khó, từ chuyên môn
- Giải thích ngắn gọn: [từ] = [nghĩa tiếng Việt]
- Sắp xếp theo thứ tự xuất hiện
- Tối đa 8-10 từ quan trọng nhất

Ví dụ format:
• [từ quan trọng]: [nghĩa tiếng Việt]"""

                print(f"[Vocab AI] Analyzing vocabulary for: {text[:50]}...")
                response = self.gemini_model.generate_content(prompt)
                explanation = response.text.strip()
                
                if explanation:
                    return f"Từ vựng quan trọng:\n\n{explanation}"
                else:
                    return "Không thể phân tích từ vựng"
                    
            except Exception as e:
                error_msg = str(e)
                print(f"[Vocab AI] Error: {error_msg}")
                
                # Handle API error and try switching keys
                error_type = self._handle_api_error(error_msg)
                
                if error_type == "switched":
                    retry_count += 1
                    print(f"[Vocab AI] Retrying with new key (attempt {retry_count}/{max_retries})")
                    continue  # Retry with new key
                elif error_type == "exhausted":
                    return "⚠️ Tất cả API keys đã hết - Không thể phân tích từ vựng"
                elif "api" in error_msg.lower() and "key" in error_msg.lower():
                    return "⚠️ API key không hợp lệ - Không thể phân tích từ vựng"
                elif "network" in error_msg.lower() or "connection" in error_msg.lower():
                    return "⚠️ Lỗi kết nối - Không thể phân tích từ vựng"
                else:
                    return f"⚠️ Lỗi phân tích từ vựng: {type(e).__name__}"
        
        # If all retries exhausted
        return "⚠️ Không thể phân tích từ vựng"
    
    def _clear_translator_ui(self):
        """Xóa tất cả nội dung trong UI dịch văn bản"""
        self.translate_input.delete("1.0", tk.END)
        self.translate_output.config(state=tk.NORMAL)
        self.translate_output.delete("1.0", tk.END)
        self.translate_output.config(state=tk.DISABLED)
        self.vocab_explanation.config(state=tk.NORMAL)
        self.vocab_explanation.delete("1.0", tk.END)
        self.vocab_explanation.config(state=tk.DISABLED)
    
    def _on_input_change(self, event=None):
        """Auto-resize input text area based on content"""
        # Get current content
        content = self.translate_input.get("1.0", tk.END)
        lines = content.count('\n')
        
        # Calculate new height (min 3, max 15)
        new_height = max(3, min(15, lines + 1))
        
        # Update height if changed
        if new_height != self.translate_input.cget('height'):
            self.translate_input.config(height=new_height)
    
    def _toggle_context_input(self):
        """Toggle context input visibility"""
        if self.use_context_var.get():
            self.context_text.pack(fill=tk.X, pady=(4, 6))
        else:
            self.context_text.pack_forget()
    


def main():
    root = tk.Tk()
    app = CambridgeDictionaryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
